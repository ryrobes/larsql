"""
Pytest fixtures and configuration for SQL connections integration tests.

Provides fixtures for:
- Docker-based services (MinIO, MongoDB, PostgreSQL, MySQL)
- Cloud services (GCS, BigQuery, S3)
- Test data and connection configuration
"""

import os
import socket
import pytest
from pathlib import Path
from typing import Dict, Any, Optional

import duckdb


# =============================================================================
# Docker Service Detection
# =============================================================================

def is_port_open(host: str, port: int, timeout: float = 1.0) -> bool:
    """Check if a port is accepting connections."""
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(timeout)
    try:
        sock.connect((host, port))
        return True
    except (socket.timeout, socket.error):
        return False
    finally:
        sock.close()


def is_minio_available() -> bool:
    """Check if MinIO is running on localhost:9100."""
    return is_port_open("localhost", 9100)


def is_mongodb_available() -> bool:
    """Check if MongoDB is running on localhost:27117."""
    return is_port_open("localhost", 27117)


def is_postgres_available() -> bool:
    """Check if PostgreSQL is running on localhost:5532."""
    return is_port_open("localhost", 5532)


def is_mysql_available() -> bool:
    """Check if MySQL is running on localhost:3406."""
    return is_port_open("localhost", 3406)


def is_cassandra_available() -> bool:
    """Check if Cassandra is running on localhost:9142."""
    return is_port_open("localhost", 9142, timeout=2.0)


# =============================================================================
# Cloud Credential Detection
# =============================================================================

def has_gcp_credentials() -> bool:
    """Check if GCP credentials are available."""
    return bool(os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"))


def has_aws_credentials() -> bool:
    """Check if AWS credentials are available."""
    return bool(
        os.environ.get("AWS_ACCESS_KEY_ID") and
        os.environ.get("AWS_SECRET_ACCESS_KEY")
    )


# =============================================================================
# Pytest Fixtures
# =============================================================================

@pytest.fixture(scope="session")
def docker_services_available() -> Dict[str, bool]:
    """Check which docker services are available."""
    return {
        "minio": is_minio_available(),
        "mongodb": is_mongodb_available(),
        "postgres": is_postgres_available(),
        "mysql": is_mysql_available(),
        "cassandra": is_cassandra_available(),
    }


@pytest.fixture(scope="session")
def cloud_credentials_available() -> Dict[str, bool]:
    """Check which cloud credentials are available."""
    return {
        "gcp": has_gcp_credentials(),
        "aws": has_aws_credentials(),
    }


@pytest.fixture(scope="function")
def duckdb_conn():
    """Create a fresh in-memory DuckDB connection for each test."""
    conn = duckdb.connect(":memory:")
    yield conn
    conn.close()


@pytest.fixture(scope="session")
def minio_config() -> Dict[str, str]:
    """MinIO connection configuration."""
    return {
        "endpoint_url": "http://localhost:9100",
        "access_key": "minioadmin",
        "secret_key": "minioadmin",
        "bucket": "test-data",
        "region": "us-east-1",
    }


@pytest.fixture(scope="session")
def mongodb_config() -> Dict[str, str]:
    """MongoDB connection configuration."""
    return {
        "uri": "mongodb://admin:testpassword@localhost:27117",
        "database": "testdb",
    }


@pytest.fixture(scope="session")
def postgres_config() -> Dict[str, str]:
    """PostgreSQL connection configuration."""
    return {
        "host": "localhost",
        "port": "5532",
        "database": "testdb",
        "user": "testuser",
        "password": "testpassword",
    }


@pytest.fixture(scope="session")
def mysql_config() -> Dict[str, str]:
    """MySQL connection configuration."""
    return {
        "host": "localhost",
        "port": "3406",
        "database": "testdb",
        "user": "testuser",
        "password": "testpassword",
    }


@pytest.fixture(scope="session")
def test_excel_file(tmp_path_factory) -> Path:
    """Create a test Excel file for excel connector tests."""
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        pytest.skip("openpyxl not installed")

    # Create temp directory that persists for the session
    tmp_dir = tmp_path_factory.mktemp("excel_data")
    excel_path = tmp_dir / "test_data.xlsx"

    # Create workbook with test data
    wb = Workbook()

    # Sheet 1: Sales
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["order_id", "product", "quantity", "price"])
    ws1.append([1001, "Widget A", 2, 29.99])
    ws1.append([1002, "Widget B", 1, 49.99])
    ws1.append([1003, "Gadget X", 3, 15.00])
    ws1.append([1004, "Widget A", 1, 29.99])
    ws1.append([1005, "Service Y", 1, 199.99])

    # Sheet 2: Customers
    ws2 = wb.create_sheet("Customers")
    ws2.append(["id", "name", "email", "tier"])
    ws2.append([101, "Alice", "alice@example.com", "gold"])
    ws2.append([102, "Bob", "bob@example.com", "silver"])
    ws2.append([103, "Carol", "carol@example.com", "bronze"])

    # Sheet 3: Products
    ws3 = wb.create_sheet("Products")
    ws3.append(["sku", "name", "price", "in_stock"])
    ws3.append(["WIDGET-A", "Widget A", 29.99, True])
    ws3.append(["WIDGET-B", "Widget B", 49.99, True])
    ws3.append(["GADGET-X", "Gadget X", 15.00, True])
    ws3.append(["SERVICE-Y", "Service Y", 199.99, False])

    wb.save(excel_path)
    return excel_path


# =============================================================================
# Skip Decorators (for use in test files)
# =============================================================================

skip_unless_minio = pytest.mark.skipif(
    not is_minio_available(),
    reason="MinIO not available - start with: docker compose -f docker-compose.sql-tests.yml up -d"
)

skip_unless_mongodb = pytest.mark.skipif(
    not is_mongodb_available(),
    reason="MongoDB not available - start with: docker compose -f docker-compose.sql-tests.yml up -d"
)

skip_unless_postgres = pytest.mark.skipif(
    not is_postgres_available(),
    reason="PostgreSQL not available - start with: docker compose -f docker-compose.sql-tests.yml up -d"
)

skip_unless_mysql = pytest.mark.skipif(
    not is_mysql_available(),
    reason="MySQL not available - start with: docker compose -f docker-compose.sql-tests.yml up -d"
)

skip_unless_cassandra = pytest.mark.skipif(
    not is_cassandra_available(),
    reason="Cassandra not available - start with: docker compose -f docker-compose.sql-tests.yml --profile cassandra up -d"
)

skip_unless_gcp = pytest.mark.skipif(
    not has_gcp_credentials(),
    reason="GCP credentials not available - set GOOGLE_APPLICATION_CREDENTIALS"
)

skip_unless_aws = pytest.mark.skipif(
    not has_aws_credentials(),
    reason="AWS credentials not available - set AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY"
)
